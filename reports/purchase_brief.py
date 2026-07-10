"""
reports/purchase_brief.py — Фаза 7: закупочный блок.

По категории превращает тренд-аналитику в решение о закупке:
  • корзины действий по стадиям диффузии:
      ПИЛОТ    (ранние последователи)  → тестовая партия 30–50 ед/цветомодель
      МАСШТАБ  (раннее большинство)    → масштабирование с модификациями
      БАЗА     (позднее большинство)   → коммерческая база, конкуренция ценой
      ИЗБЕГАТЬ (спад)                  → не входить / распродажа
  • для каждого тренда: обоснование цифрами (доля в категории, Q, рост,
    конкуренция, выручка топов), рекомендуемые цвета и материалы
    (co-occurrence по предметам категории), эскизы и ТЗ Студии.

Выход: раздел в PDF-отчёте категории + Excel для конструктора/фабрики.
Модуль вызывается из reports.trend_report; отдельный CLI:
    python -m reports.purchase_brief --category "Knitwear/Cardigan"
"""

import json
import re
import sys
from collections import Counter
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import config
import db
import taxonomy

# Корзины действий: стадия → (ключ, заголовок, действие)
BUCKETS = [
    ("РАННИЕ ПОСЛЕДОВАТЕЛИ", "ПИЛОТ",
     "Тестовая партия 30–50 ед. на цветомодель; узкий размерный ряд; "
     "решение о масштабе — по sell-through за 3–4 недели."),
    ("РАННЕЕ БОЛЬШИНСТВО", "МАСШТАБ",
     "Масштабирование только с модификациями (цвет/материал/деталь); "
     "посадка и носибельность важнее подиумной сложности."),
    ("ПОЗДНЕЕ БОЛЬШИНСТВО", "БАЗА",
     "Коммерческая база ассортимента: проверенные лекала, конкуренция "
     "ценой и цветовой линейкой, без экспериментов."),
    ("СПАД", "ИЗБЕГАТЬ",
     "В производство не входить; остатки — в распродажу."),
]
PER_BUCKET = {"ПИЛОТ": 4, "МАСШТАБ": 4, "БАЗА": 3, "ИЗБЕГАТЬ": 6}

_ARRAY_FIELDS = ("materials", "silhouette", "construction", "decoration", "colors")


def _items_with_element(conn, category, field, element):
    """SQL-условие выборки предметов категории, несущих элемент тренда."""
    base = "SELECT i.* FROM items i WHERE i.category=? AND i.confidence>=?"
    args = [category, config.MIN_ITEM_CONFIDENCE]
    if field == "pattern":
        return conn.execute(base + " AND i.pattern=?", args + [element])
    if field in _ARRAY_FIELDS:
        return conn.execute(base + f" AND i.{field} LIKE ?",
                            args + [f'%"{element}"%'])
    if field == "styles":
        return conn.execute(
            """SELECT i.* FROM items i JOIN looks l USING (look_id)
               WHERE i.category=? AND i.confidence>=? AND l.style_tags LIKE ?""",
            args + [f'%"{element}"%'])
    return conn.execute(base, args)  # category-тренд: вся категория


def _top_cooccur(conn, category, field, element, want, n=3, skip=None):
    """Топ-значения поля want у предметов категории с данным элементом."""
    cnt = Counter()
    for r in _items_with_element(conn, category, field, element):
        vals = json.loads(r[want] or "[]") if want in _ARRAY_FIELDS else [r[want]]
        for v in vals:
            if v and v != "Other" and v != (skip or element):
                cnt[v] += 1
    return [v for v, _ in cnt.most_common(n)]


def _designs(conn, trend_id):
    try:
        rows = conn.execute(
            """SELECT design_id, image_path, tech_spec, model FROM designs
               WHERE trend_id=? AND status='ok' ORDER BY design_id DESC LIMIT 4""",
            (trend_id,)).fetchall()
    except Exception:
        return []
    return [dict(r) for r in rows]


def _refs(conn, category, field, element, n=2):
    """Подиумные референсы: луки категории с элементом, свежие сезоны первыми."""
    urls, seen = [], set()
    q = _items_with_element(conn, category, field, element)
    ids = [r["look_id"] for r in q]
    if not ids:
        return []
    ph = ",".join("?" * min(len(ids), 400))
    for r in conn.execute(
            f"""SELECT image_url, designer, season_label FROM looks
                WHERE look_id IN ({ph}) ORDER BY season_year DESC""",
            ids[:400]):
        if r["image_url"] in seen:
            continue
        seen.add(r["image_url"])
        urls.append(dict(r))
        if len(urls) >= n:
            break
    return urls


def build_brief_data(conn, category, element_trends):
    """Корзины закупочного брифа из трендов категории (см. trend_report)."""
    out = []
    for stage, bucket, action in BUCKETS:
        group = [t for t in element_trends if t["stage"] == stage]
        # значимость: доля в категории, при равенстве — частотность WB
        group.sort(key=lambda t: (-t["cat_share"], -(t["q_freq"] or 0)))
        selected = group[:PER_BUCKET[bucket]]
        # Тренды с готовыми эскизами Студии включаем всегда (до +2 сверх лимита)
        extra = [t for t in group[PER_BUCKET[bucket]:]
                 if _designs(conn, t["trend_id"])][:2]
        rows = []
        for t in selected + extra:
            f, el = t["field"], t["element"]
            colors = (_top_cooccur(conn, category, f, el, "colors")
                      if f != "colors" else [el])
            materials = (_top_cooccur(conn, category, f, el, "materials")
                         if f != "materials" else [el])
            rows.append({
                "trend_id": t["trend_id"],
                "name": t["name_ru"],
                "dimension": t["type_dimension"],
                "cat_share": t["cat_share"],
                "cat_count": t["cat_count"],
                "q_freq": t["q_freq"],
                "q_growth": t["q_growth"],
                "c_cards": t["c_cards"],
                "c_top_revenue": t["c_top_revenue"],
                "colors": colors,
                "materials": materials,
                "designs": _designs(conn, t["trend_id"]),
                "refs": _refs(conn, category, f, el),
            })
        if rows:
            out.append({"bucket": bucket, "stage": stage,
                        "action": action, "trends": rows})
    return out


# ─── PDF-раздел (вызывается из trend_report) ─────────────────────────────────

def _why(t):
    parts = [f"{t['cat_count']} предм. ({t['cat_share']:.0%} категории)"]
    if t["q_freq"]:
        g = f", {t['q_growth']:+.0%} м/м" if t["q_growth"] is not None else ""
        parts.append(f"Q={t['q_freq']:,.0f}/мес{g}".replace(",", " "))
    if t["c_cards"] is not None:
        parts.append(f"карточек: {t['c_cards']:,.0f}".replace(",", " "))
    if t["c_top_revenue"]:
        parts.append(f"топ-20: {t['c_top_revenue'] / 1e6:.1f} млн ₽/30дн")
    return " · ".join(parts)


def append_brief_section(story, brief, cat_ru):
    from reportlab.lib.units import cm
    from reportlab.platypus import (KeepTogether, Paragraph, Spacer, Table,
                                    TableStyle)
    from reports import pdf_kit as K

    K.section_header(story, "Раздел 5", f"Закупочный бриф: {cat_ru}")
    story.append(Paragraph(
        "Действия по корзинам стадий диффузии. Цвета и материалы — из "
        "co-occurrence на подиуме внутри категории; ТЗ моделей — в Excel-брифе.",
        K.ST_SMALL))

    for grp in brief:
        sc = K.STAGE_HEX.get(grp["stage"], "#8a8377")
        hdr = Table([[Paragraph(
            f"<b>{grp['bucket']}</b> · {grp['stage'].lower()} — {grp['action']}",
            K.sty(size=8.5, color=K.C_WHITE, leading=11))]], colWidths=[K.INNER])
        hdr.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), K.colors.HexColor(sc)),
            ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING", (0, 0), (-1, -1), 10)]))
        blocks = [hdr, Spacer(1, 2)]
        for t in grp["trends"]:
            colors = ", ".join(taxonomy.ru("colors", c) for c in t["colors"]) or "—"
            mats = ", ".join(taxonomy.ru("materials", m) for m in t["materials"]) or "—"
            lines = [f"<b>{t['name']}</b> ({t['dimension']}) — {_why(t)}",
                     f'<font size="7">цвета: {colors} · материалы: {mats}</font>']
            if t["designs"]:
                lines.append(f'<font size="7" color="#9c4a3f">Студия: '
                             f'{len(t["designs"])} эскизов + ТЗ (Excel)</font>')
            blk = Table([[Paragraph("<br/>".join(lines), K.sty(size=8, leading=11))]],
                        colWidths=[K.INNER])
            blk.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), K.C_SURF),
                ("BOX", (0, 0), (-1, -1), 0.4, K.C_LINE),
                ("LINEBEFORE", (0, 0), (0, -1), 2, K.colors.HexColor(sc)),
                ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("LEFTPADDING", (0, 0), (-1, -1), 9)]))
            blocks += [blk, Spacer(1, 0.08 * cm)]
        story.append(KeepTogether(blocks[:4]))   # заголовок + первый тренд вместе
        for b in blocks[4:]:
            story.append(b)
        story.append(Spacer(1, 0.2 * cm))


# ─── Excel для конструктора/фабрики ──────────────────────────────────────────

def export_excel(brief, category, out_dir):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    BUCKET_FILL = {"ПИЛОТ": "2e6b4f", "МАСШТАБ": "8a6d1f",
                   "БАЗА": "4a5a7a", "ИЗБЕГАТЬ": "8a3a33"}

    wb = Workbook()

    # ── Лист 1: Закупка ──
    ws = wb.active
    ws.title = "Закупка"
    head = ["Корзина", "Действие", "Тренд", "Измерение", "Доля в категории",
            "Q WB/мес", "Рост Q м/м", "Карточек", "Выручка топ-20, ₽/30дн",
            "Цвета (подиум)", "Материалы (подиум)", "Эскизы Студии",
            "Референсы подиума"]
    ws.append(head)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1a1a1a")
        c.alignment = Alignment(wrap_text=True, vertical="center")
    for grp in brief:
        fill = PatternFill("solid", fgColor=BUCKET_FILL.get(grp["bucket"], "444444"))
        for t in grp["trends"]:
            ws.append([
                grp["bucket"], grp["action"], t["name"], t["dimension"],
                round(t["cat_share"], 3),
                int(t["q_freq"]) if t["q_freq"] else None,
                round(t["q_growth"], 3) if t["q_growth"] is not None else None,
                int(t["c_cards"]) if t["c_cards"] is not None else None,
                int(t["c_top_revenue"]) if t["c_top_revenue"] else None,
                ", ".join(taxonomy.ru("colors", c) for c in t["colors"]),
                ", ".join(taxonomy.ru("materials", m) for m in t["materials"]),
                "; ".join(d["image_path"] for d in t["designs"]) or "—",
                "\n".join(f"{r['designer']} {r['season_label']}: {r['image_url']}"
                          for r in t["refs"]) or "—",
            ])
            ws.cell(ws.max_row, 1).fill = fill
            ws.cell(ws.max_row, 1).font = Font(bold=True, color="FFFFFF")
            for col in (2, 10, 11, 12, 13):
                ws.cell(ws.max_row, col).alignment = Alignment(wrap_text=True,
                                                               vertical="top")
        ws.cell(ws.max_row, 5).number_format = "0.0%"
    for i, w in enumerate([11, 40, 24, 11, 12, 10, 10, 10, 14, 26, 26, 34, 46], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # ── Лист 2: ТЗ Студии ──
    ws2 = wb.create_sheet("ТЗ Студии")
    ws2.append(["Тренд", "Эскиз (файл)", "Модель генерации", "ТЗ конструктору"])
    for c in ws2[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1a1a1a")
    for grp in brief:
        for t in grp["trends"]:
            for d in t["designs"]:
                spec = d["tech_spec"] or ""
                try:  # JSON → читаемый текст
                    sj = json.loads(spec)
                    spec = "\n".join(f"{k}: {v}" for k, v in sj.items())
                except Exception:
                    pass
                ws2.append([t["name"], d["image_path"], d["model"], spec])
                ws2.cell(ws2.max_row, 4).alignment = Alignment(wrap_text=True,
                                                               vertical="top")
    for i, w in enumerate([24, 44, 10, 90], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    slug = re.sub(r"[^a-z0-9]+", "-", category.lower()).strip("-")
    out = Path(out_dir) / f"purchase_brief_{slug}_{date.today().isoformat()}.xlsx"
    wb.save(out)
    return out


# ─── CLI ─────────────────────────────────────────────────────────────────────

def main():
    import argparse
    from reports.trend_report import category_trends

    ap = argparse.ArgumentParser(description="Закупочный бриф (Excel) по категории")
    ap.add_argument("--category", required=True)
    ap.add_argument("--db", default=None)
    args = ap.parse_args()

    conn = db.connect(args.db or config.DB_PATH)
    trends = category_trends(conn, args.category)
    element_trends = [t for t in trends if not t["is_category_trend"]]
    brief = build_brief_data(conn, args.category, element_trends)
    if not brief:
        sys.exit("Нет трендов со стадиями для этой категории")
    out = export_excel(brief, args.category, config.REPORTS_DIR)
    print(f"Excel: {out}")


if __name__ == "__main__":
    main()
